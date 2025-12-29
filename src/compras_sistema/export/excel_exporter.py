import polars as pl
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import structlog

logger = structlog.get_logger(__name__)


class ExcelExporter:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def exportar_sugestao(self, df: pl.DataFrame, filename: str = None):
        if filename is None:
            data_hoje = datetime.now().strftime("%Y%m%d_%H%M")
            filename = f"sugestao_compras_{data_hoje}.xlsx"
        
        filepath = self.output_dir / filename
        logger.info("iniciando_export_excel", path=str(filepath))
        
        # ======== ORDEM DAS COLUNAS (REORGANIZADA) ========
        cols_export = [
            "status_diagnostico",
            "cod_produto",
            "descricao",
            "ref_fornecedor",
            "marca",
            "curva_abc",
            "curva_xyz",
            "sugestao_final",           # Sugestão Final (Resultado)
            
            # --- BLOCO DE ANÁLISE DO CÁLCULO ---
            "sugestao_calculada",       # 1. Matemática Pura
            "alerta_dados",             # 2. Auditoria (Posição Solicitada)
            "calculado_mas_bloqueado",  # 3. Flag de Bloqueio
            "motivo_bloqueio",          # 4. Razão
            # -----------------------------------
            
            "meta_pos_compra",
            "fator_sazonal",
            "lote_economico",
            "subtotal",
            "saldo_estoque",
            "saldo_oc",
            "cobertura_virtual_meses",
            "media_venda_base",
            "media_venda_dia",
            "tendencia_vendas",
            "tendencia_clientes",
            "perfil_cliente",
            "validacao_giro",
            "custo_unitario",
            "score"
        ]
        
        cols_presentes = [c for c in cols_export if c in df.columns]
        records = df.select(cols_presentes).to_dicts()
        
        # ======== CRIAÇÃO DO EXCEL ========
        wb = Workbook()
        ws = wb.active
        ws.title = "Analise Compras"
        
        # Estilos Gerais
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        center_align = Alignment(horizontal="center")
        left_align = Alignment(horizontal="left")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        # Cores
        fill_green = PatternFill(start_color="CCFFCC", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFFFE0", fill_type="solid")
        fill_orange = PatternFill(start_color="FFD700", fill_type="solid")
        fill_blue_light = PatternFill(start_color="E6F3FF", fill_type="solid")
        fill_implanta = PatternFill(start_color="E0FFFF", fill_type="solid")
        fill_red = PatternFill(start_color="FFB6C1", fill_type="solid")
        fill_alert = PatternFill(start_color="FFFF00", fill_type="solid") # Amarelo forte
        
        # Mapeamento de nomes
        mapa_nomes = {
            "ALERTA_DADOS": "⚠️ ALERTA DADOS",
            "META_POS_COMPRA": "POSIÇÃO FINAL",
            "FATOR_SAZONAL": "IDX SAZONAL",
            "MEDIA_VENDA_DIA": "GIRO DIA (AJUST)",
            "MEDIA_VENDA_BASE": "GIRO DIA (BASE)",
            "COBERTURA_VIRTUAL_MESES": "COBERTURA MESES",
            "REF_FORNECEDOR": "REF. FABRICA",
            "SUGESTAO_CALCULADA": "CALC. ORIGINAL",
            "CALCULADO_MAS_BLOQUEADO": "BLOQUEADO?",
            "MOTIVO_BLOQUEIO": "MOTIVO BLOQUEIO"
        }
        
        headers = [c.replace("_", " ").upper() for c in cols_presentes]
        headers = [mapa_nomes.get(h, h) for h in headers]
        
        ws.append(headers)
        
        # Formata cabeçalho
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        # Preenche dados
        for row_idx, row_data in enumerate(records, 2):
            for col_idx, col_name in enumerate(cols_presentes, 1):
                val = row_data[col_name]
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                
                # Alinhamento
                if col_name == "descricao":
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                
                # Formatação numérica
                if col_name in ["custo_unitario", "subtotal"]:
                    cell.number_format = 'R$ #,##0.00'
                elif col_name in ["media_venda_dia", "media_venda_base", "fator_sazonal"]:
                    cell.number_format = '0.00'
                elif col_name in ["cobertura_virtual_meses"]:
                    cell.number_format = '0.0'
                elif col_name == "score":
                    cell.number_format = '#,##0'
                
                # ======== FORMATAÇÃO CONDICIONAL ========
                
                # 1. ALERTA DE DADOS
                if col_name == "alerta_dados" and val:
                    cell.font = Font(bold=True, color="FF0000")
                    cell.fill = fill_alert
                
                # 2. Sugestões de compra
                if col_name in ["sugestao_final", "subtotal"] and row_data.get("sugestao_final", 0) > 0:
                    cell.font = Font(bold=True, color="006400")
                    cell.fill = fill_green
                
                # 3. Produtos bloqueados
                if col_name == "calculado_mas_bloqueado" and val == "SIM":
                    cell.fill = fill_red
                    cell.font = Font(bold=True, color="8B0000")
                
                # 4. Motivo do bloqueio
                if col_name == "motivo_bloqueio" and val:
                    cell.font = Font(color="DC143C", italic=True)
                
                # 5. Fator sazonal
                if col_name == "fator_sazonal":
                    if isinstance(val, (int, float)):
                        if val < 0.90:
                            cell.font = Font(color="0000FF")
                            cell.fill = fill_blue_light
                        elif val > 1.10:
                            cell.font = Font(color="B22222", bold=True)
                
                # 6. Status diagnóstico
                if col_name == "status_diagnostico":
                    val_str = str(val).upper()
                    if "IMPLANTAÇÃO" in val_str:
                        cell.fill = fill_implanta
                        cell.font = Font(color="00008B", bold=True)
                    elif "RUPTURA" in val_str:
                        cell.fill = PatternFill(start_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif "BLOQUEADO" in val_str:
                        cell.fill = PatternFill(start_color="808080", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif "INATIVO" in val_str:
                        cell.fill = PatternFill(start_color="000000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif "ALERTA" in val_str:
                        cell.fill = PatternFill(start_color="FF8C00", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif "EXCESSO" in val_str:
                        cell.fill = fill_yellow
                    elif "COMPRAR" in val_str:
                        cell.fill = fill_green
                
                # 7. Tendência
                if col_name == "tendencia_vendas":
                    val_str = str(val).upper()
                    if "ALTA" in val_str:
                        cell.font = Font(color="006400", bold=True)
                    elif "QUEDA" in val_str:
                        cell.font = Font(color="FF0000", bold=True)
                
                # 8. Validação Giro
                if col_name == "validacao_giro":
                    val_str = str(val)
                    if "ITEM NOVO" in val_str:
                        cell.fill = fill_implanta
                        cell.font = Font(color="00008B", bold=True)
                    elif "SEM MOVIMENTO" in val_str:
                        cell.font = Font(color="808080", italic=True)
                    elif "Excesso" in val_str:
                        cell.font = Font(bold=True, color="B22222")
                        cell.fill = fill_orange
        
        # Ajuste de largura
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(col_idx)
            limit = 60 if cols_presentes[col_idx-1] == "descricao" else 40
            
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            ws.column_dimensions[column].width = min(max_length + 3, limit)
        
        wb.save(filepath)
        logger.info("export_excel_concluido")
        return filepath